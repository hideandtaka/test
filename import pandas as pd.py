import pandas as pd

# サンプルの入力Excelを作成
df_input = pd.DataFrame([
    ("りんご", 10, 120),
    ("バナナ", 5, 80),
    ("みかん", 12, 60),
], columns=["商品名", "数量", "単価"])

df_input.to_excel("sales_input.xlsx", index=False)
print("sales_input.xlsx を作成しました")
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def calculate_daily_sales(sales):
    return sum(qty * price for _, qty, price in sales)

# --- Excelから読み込み ---
df = pd.read_excel("sales_input.xlsx")
sales = list(df.itertuples(index=False, name=None))  # リスト of タプルに変換
total = calculate_daily_sales(sales)

# --- 結果をExcelに出力 ---
df["小計"] = df["数量"] * df["単価"]
df.loc["合計"] = ["合計", "", "", df["小計"].sum()]

with pd.ExcelWriter("sales_output.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="売上")
    
    ws = writer.sheets["売上"]
    
    # ヘッダーを太字・背景色
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", start_color="4472C4", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # 列幅を自動調整
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

print(f"本日の売上合計: {total} 円")
print("sales_output.xlsx に出力しました")
